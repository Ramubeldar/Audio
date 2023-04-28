# program for setting title as well as text on video

# importing important python libraries
import moviepy.editor as mp
from openpyxl import load_workbook
from moviepy.editor import *
import shutil
from gtts import gTTS
import os


# defining the function
def read_excel_row(row_num, excel_file):
    global font_first, font_second
    wb = load_workbook(filename=excel_file, read_only=True)
    worksheet = wb.active

    # row data preprocessing
    Font_Code = []
    for col in range(6, 7):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Font_Code.append(cell_value)
    # Split the single string element into a list of separate elements
    Font_Code = Font_Code[0].split('\n')

    # Create a new list to hold the elements without newline character
    Font_Codes = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Font_Code:
        Font_Codes.append(element.strip())

    Parent_folder_name = []
    for col in range(1, 2):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Parent_folder_name.append(cell_value)

    Input_File_Name = []
    for col in range(2, 3):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Input_File_Name.append(cell_value)

    Message = []
    for col in range(3, 4):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Message.append(cell_value)

    Duration = []
    for col in range(4, 5):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Duration.append(cell_value)

    Output_File_Name = []
    for col in range(5, 6):
        cell_value = worksheet.cell(row=row_num, column=col).value
        Output_File_Name.append(cell_value)

    Parent_folder_name = Parent_folder_name[0].split('\n')

    # Create a new list to hold the elements without newline character
    Parent_folder_names = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Parent_folder_name:
        Parent_folder_names.append(element.strip())

    Input_File_Name = Input_File_Name[0].split('\n')

    # Create a new list to hold the elements without newline character
    Input_File_Names = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Input_File_Name:
        Input_File_Names.append(element.strip())

    # Split the single string element into a list of separate elements
    Message = Message[0].split('\n')

    # Create a new list to hold the elements without newline character
    Messages = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Message:
        Messages.append(element.strip())

    # Split the single string element into a list of separate elements
    Duration = Duration[0].split('\n')

    # Create a new list to hold the elements without newline character
    Durations = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Duration:
        Durations.append(element.strip())

    Output_File_Name = Output_File_Name[0].split('\n')

    # Create a new list to hold the elements without newline character
    Output_File_Names = []

    # Loop through the original list, strip leading and trailing whitespaces and append the elements to the new list
    for element in Output_File_Name:
        Output_File_Names.append(element.strip())
    try:
        video = mp.VideoFileClip(Input_File_Names[0])
    except:
        print("video file is not available for the row number: - ", row_num, Parent_folder_names[0])
        with open('error_log.txt', 'a') as f:
            data = str(Input_File_Names[0]) + "--->video file is not available for the row_number--->" + str(row_num)
            f.write(data + "\n")
        return

    video_sub = video
    clips = []
    head_i = 0
    message_iterate = 0
    time_iterate = 0

    while message_iterate < len(Message):

        text = Messages[message_iterate]

        # define function to generate audio for a subtitle message

        def generate_audio(message,speed=1.1):   #,speed=1.1
            if Font_Codes[0] == 'Arabic':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='ar')
            elif Font_Codes[0] == 'Afrikaans':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='af')
            elif Font_Codes[0] == 'Korean':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='ko')
            elif Font_Codes[0] == 'Japanese':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='ja')
            elif Font_Codes[0] == 'Italian':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='it')
            elif Font_Codes[0] == 'German':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='de')
            elif Font_Codes[0] == 'French':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='fr')
            elif Font_Codes[0] == 'Swahili':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='sw')
            elif Font_Codes[0] == 'Portuguese':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='pt')
            elif Font_Codes[0] == 'Malay':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='ms')
            elif Font_Codes[0] == 'Spanish':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='es')
            elif Font_Codes[0] == 'Hindi':

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='hi')

            else:

                # create GTTS object and specify language
                tts = gTTS(text=message, lang='en')

            tts.save('audio.mp3')
            # load audio file and return audio clip
            return AudioFileClip('audio.mp3')

        message_iterate += 1
        # formatting time
        start_time = Durations[time_iterate]
        time_iterate += 1
        hours, minutes, seconds, milliseconds = map(int, start_time.split(':'))
        start_time = (hours * 3600) + (minutes * 60) + seconds + (milliseconds / 1000)
        end_time = start_time

        bg_color = '#FFC10C'
        color = 'black'
        margin = 50

        if Font_Codes[0] == 'AR':  # arabic
            font_first = 'Expo Arabic Bold'
            font_second = 'Times New Roman'

        elif Font_Codes[0] == 'KO' or 'JA':  # korean or japanes
            font_first = 'gulim'
            font_second = 'gulim'
        elif Font_Codes[0] == 'Hindi':  # korean or japanes
            font_first = 'KRDEV011'
            font_second = 'KRDEV011'

        else:  # default
            font_first = 'BerkshireSwash-Regular'
            font_second = 'Aleo-Regular'

        if head_i == 0:
            head_i += 1
            subtitle = (
                mp.TextClip(text, fontsize=100, color='white', font=font_first, stroke_color='black', stroke_width=2))
            subtitle = subtitle.set_start(start_time).set_end(end_time).set_position(
                ('center', video_sub.size[1] - subtitle.size[1] - margin))

            # generate audio clip for subtitle message
            audio_clip = generate_audio(text)
            # Increase the speed of the audio by 1.1x
            audio_clip= audio_clip.fx(vfx.speedx, factor=1.1)
            # combine the audio and text clips into a single subtitle clip
            subtitle = CompositeVideoClip([subtitle.set_audio(audio_clip)])
            # clips.append(subtitle)
            clips.append(subtitle)
        else:
            subtitle = (mp.TextClip(text, fontsize=54, color=color, font=font_second,
                                    bg_color=bg_color)
                        )
            subtitle = subtitle.set_start(start_time).set_end(end_time).set_position(
                ('center', video_sub.size[1] - subtitle.size[1] - margin))

            # generate audio clip for subtitle message
            audio_clip = generate_audio(text)
            # Increase the speed of the audio by 1.1x
            audio_clip = audio_clip.fx(vfx.speedx, factor=1.2)
            # set audio clip to start at the same time as the text clip
            audio_clip = audio_clip.set_start(start_time)
            # combine the audio and text clips into a single subtitle clip
            subtitle = CompositeVideoClip([subtitle.set_audio(audio_clip)])
            clips.append(subtitle)
    # create a CompositeVideoClip with the TextClips and the original video clip
    composite_clip = mp.CompositeVideoClip([video_sub] + clips)

    # specify the name of the top-level folder
    top_level_folder_name = Parent_folder_names[0]

    # specify the name of the subfolder
    subfolder_name = Font_Codes[0]

    # check if the top-level folder already exists
    if not os.path.exists(top_level_folder_name):
        # create the top-level folder
        os.makedirs(top_level_folder_name)

    # check if the subfolder already exists
    subfolder_path = os.path.join(top_level_folder_name, subfolder_name)
    if os.path.exists(subfolder_path):
        # remove the subfolder and all its contents
        shutil.rmtree(subfolder_path)

    # create the subfolder within the top-level folder
    os.makedirs(subfolder_path)

    # Save the final edited video file in the child folder
    final_video_path = os.path.join(subfolder_path, Output_File_Names[0])

    composite_clip.write_videofile(final_video_path)
    # os.remove(filename)
    print("completed rows number ", row_num)


# reading the exel file
excel_file_sheets = "Localized Text - Recipes and Yoga.xlsx"
workbook = load_workbook("Localized Text - Recipes and Yoga.xlsx")

# Select the active worksheet
worksheet = workbook.active

# Find the number of rows with data in excel sheet
num_rows = 0
for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
    if any(cell.value for cell in row):
        num_rows += 1

for row_number in range(10, 10 + 1):
    read_excel_row(row_number, excel_file_sheets)
